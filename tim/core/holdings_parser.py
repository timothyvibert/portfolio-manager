"""Parser for UBS-style client holdings XLSX files.

The production file (work machine) has multiple tabs:
  Equities | Non Traditional | Cash | Disclaimers
The committed sample at tim/data/Holdings.xlsx may have been flattened
to a single Sheet1. The parser handles both via HOLDINGS_SHEET_CANDIDATES.

Output: a ParsedPortfolio dataclass with three DataFrames
(equity_positions, option_positions, other_positions), a portfolio_total
dict, and a parse_warnings list.

Listed options are recognized by the '99UB' CUSIP prefix and have their
market_value taken from the Value column (the brokerage already bakes the
×100 contract multiplier into Value, so we never recompute it from
quantity × price). A sanity check warns — but does not raise — if the
recomputed value diverges by more than 1% (or $0.50, whichever is
larger).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from tim.config import (
    DEFAULT_HOLDINGS_FILE,
    EXCHANGE_SUFFIX_MAP,
    HOLDINGS_SHEET_CANDIDATES,
    NON_US_CINS_PREFIXES,
    NON_US_CURRENCY_TOKENS,
    NON_US_STYLE_VALUES,
    OPTION_CONTRACT_MULTIPLIER,
)
from tim.core.ticker_utils import construct_option_ticker


# ---------------------------------------------------------------------------
# Output dataclass
# ---------------------------------------------------------------------------

@dataclass
class ParsedPortfolio:
    equity_positions: pd.DataFrame
    option_positions: pd.DataFrame
    other_positions: pd.DataFrame
    portfolio_total: dict
    parse_warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sheet selection / header detection
# ---------------------------------------------------------------------------

def _select_sheet(file_path: Path) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheetnames = list(wb.sheetnames)
    wb.close()
    for candidate in HOLDINGS_SHEET_CANDIDATES:
        if candidate in sheetnames:
            return candidate
    for sn in sheetnames:
        df = pd.read_excel(file_path, sheet_name=sn, header=None, nrows=30)
        for i in range(len(df)):
            row_vals = [str(v).strip() for v in df.iloc[i].tolist() if pd.notna(v)]
            if row_vals and row_vals[0] == "Info":
                return sn
    raise ValueError(
        f"Could not find a sheet matching {HOLDINGS_SHEET_CANDIDATES} or "
        f"containing an 'Info' header in {file_path}. Available: {sheetnames}"
    )


def _find_header_row(df: pd.DataFrame) -> int:
    for i in range(min(20, len(df))):
        row = df.iloc[i].tolist()
        non_null = [str(v).strip() for v in row if pd.notna(v)]
        if non_null and non_null[0] == "Info":
            return i
    raise ValueError("Could not find header row containing 'Info'.")


def _normalize_header(name: object) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    text = str(name).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ---------------------------------------------------------------------------
# Numeric / string coercion
# ---------------------------------------------------------------------------

def _clean_numeric(val) -> float | None:
    """Strip '*' footnote markers, '^' price prefix, commas, '%'.
    Returns None for blank, '*'-only, em-dash, or unparseable."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("", "*", "\u2014", "-"):
        return None
    s = s.lstrip("*").strip()
    s = s.lstrip("^").strip()
    s = s.replace(",", "").rstrip("%")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_text(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Position classification
# ---------------------------------------------------------------------------

def _classify_instrument(cusip: str, description: str,
                         quantity: float | None) -> str:
    cusip_u = (cusip or "").upper().strip()
    desc_u = (description or "").upper()
    if cusip_u.startswith("99UB"):
        return "option"
    if "WTS " in desc_u or "WARRANT" in desc_u:
        return "warrant"
    if quantity is not None and quantity != 0:
        return "equity"
    return "other"


_OPT_RE = re.compile(
    r"^(PUT|CALL)\s+(.+?)\s+DUE\s+(\d{2}/\d{2}/\d{2})\s+([\d.]+)\s*(.*)$",
    re.IGNORECASE,
)


def _parse_option_description(desc: str) -> dict | None:
    """Returns {right, issuer_name, expiry, strike, internal_code} or None."""
    if not desc:
        return None
    flat = re.sub(r"\s+", " ", desc.replace("\n", " ")).strip()
    m = _OPT_RE.match(flat)
    if not m:
        return None
    right = m.group(1).upper()
    issuer = m.group(2).strip()
    expiry = datetime.strptime(m.group(3), "%m/%d/%y").date()
    strike = float(m.group(4))
    internal = m.group(5).strip()
    return {
        "right": right,
        "issuer_name": issuer,
        "expiry": expiry,
        "strike": strike,
        "internal_code": internal,
    }


# ---------------------------------------------------------------------------
# Region and ticker construction
# ---------------------------------------------------------------------------

def _classify_region(symbol: str, cusip: str, style: str | None,
                     description: str | None) -> str:
    if style and style.strip() in NON_US_STYLE_VALUES:
        return "non_us"
    if description:
        upper_desc = description.upper()
        for tok in NON_US_CURRENCY_TOKENS:
            if tok.upper() in upper_desc:
                return "non_us"
    if cusip:
        first = cusip.strip()[:1].upper()
        if first and first in NON_US_CINS_PREFIXES:
            return "non_us"
    return "us"


def _underlying_bbg_ticker(symbol: str, region: str) -> str | None:
    """'AAPL US Equity' for US, 'RACE IM Equity' for mapped non-US,
    None for unmapped non-US."""
    if not symbol:
        return None
    if region == "us":
        return f"{symbol} US Equity"
    suffix = EXCHANGE_SUFFIX_MAP.get(symbol.upper())
    if suffix is None:
        return None
    return f"{symbol} {suffix} Equity"


# ---------------------------------------------------------------------------
# Sanity check
# ---------------------------------------------------------------------------

def _within_tolerance(reported: float, recomputed: float) -> bool:
    """1% of |reported| or $0.50 — whichever is larger."""
    tol = max(0.01 * abs(reported), 0.5)
    return abs(reported - recomputed) <= tol


# ---------------------------------------------------------------------------
# Main parse
# ---------------------------------------------------------------------------

def parse_holdings(file_path: Path = DEFAULT_HOLDINGS_FILE,
                   sheet_name: str | None = None) -> ParsedPortfolio:
    file_path = Path(file_path)
    if sheet_name is None:
        sheet_name = _select_sheet(file_path)

    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    header_idx = _find_header_row(raw)
    header_row = raw.iloc[header_idx].tolist()
    headers = [_normalize_header(h) for h in header_row]

    body = raw.iloc[header_idx + 1:].reset_index(drop=True)
    body.columns = headers

    required = {
        "Info", "CUSIP", "Style", "Symbol", "Description", "Quantity",
        "Price", "Value",
    }
    missing = required - set(headers)
    if missing:
        raise ValueError(
            f"Missing expected column(s) {missing} in sheet '{sheet_name}'. "
            f"Found columns: {headers}"
        )

    col_total_cost = "Total Cost"
    col_unrealized = "Unrealized (Tax) G/L"
    col_pct_pnl = "%G/L"
    col_gain_loss = "Gain-Loss/ Inv Return"
    col_avg_cost = "Avg Cost"
    col_tax_lots = "Tax Lots"

    parse_warnings: list[str] = []

    # Identify totals row: first row where Symbol is null AND Value is not null.
    portfolio_total: dict = {}
    totals_row_idx: int | None = None
    for i in range(len(body)):
        row = body.iloc[i]
        sym = _clean_text(row.get("Symbol"))
        val = _clean_numeric(row.get("Value"))
        if not sym and val is not None:
            totals_row_idx = i
            portfolio_total = {
                "total_market_value": val,
                "total_cost": _clean_numeric(row.get(col_total_cost)),
                "total_unrealized": _clean_numeric(row.get(col_unrealized)),
                "total_pct_pnl": _clean_numeric(row.get(col_pct_pnl)),
                "total_gain_loss": _clean_numeric(row.get(col_gain_loss)),
            }
            break

    if totals_row_idx is None:
        portfolio_total = {
            "total_market_value": None,
            "total_cost": None,
            "total_unrealized": None,
            "total_pct_pnl": None,
            "total_gain_loss": None,
        }
        parse_warnings.append(
            "No totals row detected (no row with empty Symbol but populated Value)."
        )
        positions = body
    else:
        positions = body.iloc[totals_row_idx + 1:].reset_index(drop=True)

    equity_rows: list[dict] = []
    option_rows: list[dict] = []
    other_rows: list[dict] = []

    for _, row in positions.iterrows():
        cusip = _clean_text(row.get("CUSIP"))
        symbol = _clean_text(row.get("Symbol"))
        description = _clean_text(row.get("Description"))
        style = _clean_text(row.get("Style")) or None
        quantity = _clean_numeric(row.get("Quantity"))
        price = _clean_numeric(row.get("Price"))
        market_value = _clean_numeric(row.get("Value"))
        cost_basis = _clean_numeric(row.get(col_total_cost))
        unrealized = _clean_numeric(row.get(col_unrealized))
        pct_pnl = _clean_numeric(row.get(col_pct_pnl))
        avg_cost = _clean_numeric(row.get(col_avg_cost))
        tax_lots = _clean_numeric(row.get(col_tax_lots))

        # Skip blank rows entirely
        if not cusip and not symbol and not description and quantity is None:
            continue

        kind = _classify_instrument(cusip, description, quantity)
        region = _classify_region(symbol, cusip, style, description)
        underlying_ticker = _underlying_bbg_ticker(symbol, region)

        if kind == "option":
            opt = _parse_option_description(description)
            if opt is None:
                parse_warnings.append(
                    f"Could not parse option description for CUSIP {cusip} "
                    f"(symbol {symbol!r}); routing to other_positions."
                )
                other_rows.append({
                    "cusip": cusip,
                    "symbol": symbol,
                    "description": description,
                    "quantity": quantity,
                    "market_value": market_value,
                    "manual_review_reason": "unparseable_option_description",
                })
                continue

            if underlying_ticker is None:
                parse_warnings.append(
                    f"Non-US underlying {symbol!r} (CUSIP {cusip}) has no "
                    f"EXCHANGE_SUFFIX_MAP entry; option routed to "
                    f"other_positions for manual review."
                )
                other_rows.append({
                    "cusip": cusip,
                    "symbol": symbol,
                    "description": description,
                    "quantity": quantity,
                    "market_value": market_value,
                    "manual_review_reason": "non_us_unmapped_exchange",
                })
                continue

            bbg_ticker = construct_option_ticker(
                underlying_ticker, opt["expiry"], opt["right"], opt["strike"],
                sector_hint="Equity",
            )

            dte_calendar = (opt["expiry"] - datetime.now().date()).days

            currency_hint = None
            for tok in NON_US_CURRENCY_TOKENS:
                if tok.upper() in description.upper():
                    currency_hint = tok.strip()
                    break

            if (price is not None and quantity is not None and quantity != 0
                    and market_value is not None):
                recomputed = quantity * price * OPTION_CONTRACT_MULTIPLIER
                if not _within_tolerance(market_value, recomputed):
                    parse_warnings.append(
                        f"Option {bbg_ticker}: qty*price*100={recomputed:.2f} "
                        f"diverges from reported value={market_value:.2f}."
                    )

            option_rows.append({
                "bbg_ticker": bbg_ticker,
                "underlying_symbol": symbol,
                "underlying_bbg_ticker": underlying_ticker,
                "region": region,
                "right": opt["right"],
                "strike": opt["strike"],
                "expiry": opt["expiry"],
                "dte_calendar": dte_calendar,
                "currency_hint": currency_hint,
                "quantity": quantity,
                "price": price,
                "market_value": market_value,
                "cost_basis": cost_basis,
                "unrealized_pnl": unrealized,
                "pct_pnl": pct_pnl,
                "internal_code": opt["internal_code"],
                "multiplier": OPTION_CONTRACT_MULTIPLIER,
            })

        elif kind == "warrant":
            other_rows.append({
                "cusip": cusip,
                "symbol": symbol,
                "description": description,
                "quantity": quantity,
                "market_value": market_value,
                "manual_review_reason": "warrant",
            })

        elif kind == "equity":
            if (price is not None and quantity is not None
                    and market_value is not None):
                recomputed = quantity * price
                if not _within_tolerance(market_value, recomputed):
                    parse_warnings.append(
                        f"Equity {symbol}: qty*price={recomputed:.2f} "
                        f"diverges from reported value={market_value:.2f}."
                    )

            equity_rows.append({
                "bbg_ticker": underlying_ticker,
                "symbol": symbol,
                "cusip": cusip,
                "style": style,
                "region": region,
                "description": description,
                "quantity": quantity,
                "price": price,
                "market_value": market_value,
                "cost_basis": cost_basis,
                "unrealized_pnl": unrealized,
                "pct_pnl": pct_pnl,
                "avg_cost": avg_cost,
                "tax_lots": tax_lots,
                "multiplier": 1,
            })

        else:
            other_rows.append({
                "cusip": cusip,
                "symbol": symbol,
                "description": description,
                "quantity": quantity,
                "market_value": market_value,
                "manual_review_reason": "unclassified",
            })

    equity_cols = [
        "bbg_ticker", "symbol", "cusip", "style", "region", "description",
        "quantity", "price", "market_value", "cost_basis", "unrealized_pnl",
        "pct_pnl", "avg_cost", "tax_lots", "multiplier",
    ]
    option_cols = [
        "bbg_ticker", "underlying_symbol", "underlying_bbg_ticker", "region",
        "right", "strike", "expiry", "dte_calendar", "currency_hint",
        "quantity", "price", "market_value", "cost_basis", "unrealized_pnl",
        "pct_pnl", "internal_code", "multiplier",
    ]
    other_cols = [
        "cusip", "symbol", "description", "quantity", "market_value",
        "manual_review_reason",
    ]

    equity_df = pd.DataFrame(equity_rows, columns=equity_cols)
    option_df = pd.DataFrame(option_rows, columns=option_cols)
    other_df = pd.DataFrame(other_rows, columns=other_cols)

    return ParsedPortfolio(
        equity_positions=equity_df,
        option_positions=option_df,
        other_positions=other_df,
        portfolio_total=portfolio_total,
        parse_warnings=parse_warnings,
    )


def get_unique_underlyings(portfolio: ParsedPortfolio) -> list[str]:
    """Sorted union of equity_positions.bbg_ticker and
    option_positions.underlying_bbg_ticker (drops None / NaN)."""
    tickers: set[str] = set()
    if not portfolio.equity_positions.empty:
        for v in portfolio.equity_positions["bbg_ticker"].tolist():
            if isinstance(v, str) and v:
                tickers.add(v)
    if not portfolio.option_positions.empty:
        for v in portfolio.option_positions["underlying_bbg_ticker"].tolist():
            if isinstance(v, str) and v:
                tickers.add(v)
    return sorted(tickers)
